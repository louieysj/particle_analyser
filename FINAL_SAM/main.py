import os
import sys
import time

import cv2
import numpy as np
import openpyxl
import win32api
import win32con
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QLabel

import utils
from UI.main_window_ui import Ui_MainWindow
from UI.measure_scale_dialog import ActualScaleDialog
from detector import SAMDetector


class MainWindow(QMainWindow):
    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        # 1. 直接load ui:
        # loadUi(r'UI\MainWindow.ui', self)
        # 2. 先 pyuic5  UI/MainWindow.ui -o UI/main_window_ui.py, 后load
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.status_label = QLabel('')
        self.ui.statusbar.addWidget(self.status_label)
        # self.ui.slider_depth.setTracking(False)

        # ----------参数初始化-------------
        self.currentFolder = None  # 当前文件夹
        self.imgList = None  # 当前文件夹下的图片文件名
        self.formerImgIdx = None
        self.currentImgIdx = 0  # imgList的当前index
        self.currentImg = None
        self.currentImgDepth = None
        self.contours = None  # 当天打开的img的检测到的所有轮廓，原始数据
        self.show_scale = self.ui.slider_scale.value() / 100  # cv2.imshow的缩放比例
        self.actual_scale = None  # 真实世界的比例尺 um/每像素
        self.minAreaIndex = None  # 筛选面积大小
        self.currentImgAreas = None
        self.DepthThresh = None  # 筛选深度大小
        self.lineWidth = 15

        self.FLAG_modify_mode = False

        self.detector = SAMDetector()
        # 绑定按钮
        self.connectSlots()

    def connectSlots(self):
        # buttons
        self.ui.btn_goto_previous.clicked.connect(self.goto_previous)
        self.ui.btn_goto_next.clicked.connect(self.goto_next)
        self.ui.btn_detect_contour.clicked.connect(self.detect_contour)

        # others
        self.ui.slider_scale.valueChanged.connect(self.scale_OnChanged)
        # self.ui.slider_depth.valueChanged.connect(self.depth_range_changed)
        self.ui.slider_area.valueChanged.connect(self.area_range_changed)
        self.ui.checkBox_delete_edge_contour.stateChanged.connect(self.show_result)
        self.ui.checkBox_draw_contours.stateChanged.connect(self.show_result)

        # actions
        self.ui.actionOpenFolder.triggered.connect(self.OpenFolder)
        self.ui.actionModifyMode.triggered.connect(self.switchModifyMode)
        self.ui.actionMeasureScale.triggered.connect(self.measureActualScaleMod)
        self.ui.actionImportScale.triggered.connect(self.ImportScaleFromImage)
        self.ui.actionExportData.triggered.connect(self.ExportData)

        # 自定义组件
        # self.ui.RangeSlider_area.rangeChanged.connect(self.area_range_changed)

    ################################ Slots ################################
    # --------------------------pushbutoon slots---------------------------
    def detect_contour(self):
        if not self.imgList:
            QMessageBox.information(self, 'info', 'Open Folder First!')
            return

        print('detect_contour')
        self.read_img()
        st = time.time()
        # self.contours = detect_TEST(self.currentImg, self.currentImgDepth, self.DepthThresh)
        self.contours = self.detector.detect(self.currentImg)
        print(time.time() - st)

        areas = [cv2.contourArea(c) for c in self.contours]
        self.currentImgAreas = sorted(areas)
        self.minAreaIndex = int(len(areas)*0.2)
        self.ui.slider_area.setRange(0, len(areas)-1)
        self.ui.slider_area.setValue(self.minAreaIndex)

        self.show_result()
        cv2.waitKey()

    def goto_next(self):
        if not self.imgList:
            return
        self.currentImgIdx = min(len(self.imgList) - 1, self.currentImgIdx + 1)
        self.ui.label_current_file.setText(
            f'({self.currentImgIdx + 1}/{len(self.imgList)}) [{self.imgList[self.currentImgIdx]}]')

        # self.detect_contour()

    def goto_previous(self):
        if not self.imgList:
            return
        self.currentImgIdx = max(0, self.currentImgIdx - 1)
        self.ui.label_current_file.setText(
            f'({self.currentImgIdx + 1}/{len(self.imgList)}) [{self.imgList[self.currentImgIdx]}]')

        # self.detect_contour()

    # ---------------------------slider slots-------------------------------
    def scale_OnChanged(self, ):
        self.show_scale = self.ui.slider_scale.value() / 100
        self.ui.label_scale.setText(f'ShowScale:{self.show_scale:.2f}')
        if self.contours:
            self.show_result()

    def area_range_changed(self, min_area_idx):
        # print(min_area, max_area)
        if not self.contours:
            return
        self.minAreaIndex = min_area_idx
        min_area = min(self.currentImgAreas)
        max_area = max(self.currentImgAreas)
        self.ui.label_minmaxArea.setText(
            f'MinArea:{min_area:.1f}, MaxArea:{max_area:.1f}\n'
            f'Filter MinArea:{self.currentImgAreas[min_area_idx]:.1f}'
        )
        self.show_result()

    # ---------------------------action slots-------------------------------    
    def OpenFolder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select Directory")  # 打开文件夹选择对话框
        if folder_path:
            imgList = os.listdir(folder_path)
            allow_ext = ['jpg', 'png', 'tif', 'bmp']
            imgList = [i for i in imgList if i.lower().split('.')[-1] in allow_ext and not i.startswith('PREPROCESSED_')]
            if imgList:
                self.currentFolder = folder_path
                self.imgList = imgList
                self.currentImgIdx = 0
                self.formerImgIdx = None
                # self.read_img_and_depth()

                self.ui.label_current_folder.setText(self.currentFolder)
                self.ui.label_current_file.setText(
                    f'({self.currentImgIdx + 1}/{len(self.imgList)}) [{self.imgList[self.currentImgIdx]}]')
            else:
                QMessageBox.information(self, 'info', 'There is no image in the folder!')

    def ExportData(self):
        if not self.actual_scale:
            QMessageBox.information(self, 'info', 'Please set scale first!')
            return
        if self.FLAG_modify_mode:
            QMessageBox.information(self, 'info', "You haven't quit modify mode!")
            return

        print('ExportData')
        try:
            excel_filename = f'{os.path.basename(self.currentFolder)}.xlsx'
            excel_filename = os.path.join(self.currentFolder, excel_filename)
            if os.path.exists(excel_filename):
                workbook = openpyxl.load_workbook(excel_filename)
            else:
                workbook = openpyxl.Workbook()
            if self.imgList[self.currentImgIdx] in workbook.sheetnames:
                del workbook[self.imgList[self.currentImgIdx]]
                # worksheet.delete_rows(1, worksheet.max_row)
                # worksheet.delete_cols(1, worksheet.max_column)

            # 写入数据
            worksheet = workbook.create_sheet(self.imgList[self.currentImgIdx])
            worksheet.append(['Circularity', 'Area', 'Diameter'])
            diameters = []
            for contour in self.contours_filter(self.contours):
                rescaled = (contour * self.actual_scale).astype(np.float32)
                Circularity = utils.cal_circularity(rescaled)
                Diameter = utils.cal_diameter(rescaled)
                diameters.append(Diameter)
                Area = utils.cal_area(rescaled)
                worksheet.append([Circularity, Area, Diameter, ])

            # 写入比例尺
            worksheet = workbook.worksheets[0]
            worksheet.append([self.imgList[self.currentImgIdx], self.actual_scale])

            # 显示颗粒总数量和d97, d50
            total_num = len(diameters)
            sorted_diameters = sorted(diameters)
            d97 = sorted_diameters[int((total_num - 1) * 0.97)]
            d50 = sorted_diameters[int((total_num - 1) * 0.5)]
            self.ui.statusbar.showMessage(f"TotalNum:{total_num}, d97:{d97:.3f}, d50:{d50:.3f}")

            workbook.save(excel_filename)
            print('Excel saved!')
            QMessageBox.information(self, 'info', 'Saved!')
        except Exception as e:
            print(e)
            QMessageBox.warning(self, 'Error Happened!!!', str(e))

    def switchModifyMode(self, checked):
        if not self.contours:
            self.ui.actionModifyMode.setChecked(False)
            return
        print('switch to Delete Mode')
        if checked:
            QMessageBox.information(self,
                'ModifyMode',
                'Press the left button to start drawing, release it to end drawing;\n'
                'Right click to delete: If it is a hand shape, delete the contour at the current position, '
                'otherwise delete the largest contour surrounding the current position;\n'
                'Middle-click to undo'
            )

        if checked:
            self.FLAG_modify_mode = True
            # cv2.namedWindow('modify', cv2.WINDOW_NORMAL)
            cv2.setMouseCallback('result', self.mouse_callback_modify)
            self.contour_history = [self.contours]  # 临时
        else:
            self.FLAG_modify_mode = False
            # cv2.destroyWindow('modify')
            cv2.setMouseCallback('result', lambda *args: None)
            self.contours = self.contour_history[-1]
        self.show_result()

    def measureActualScaleMod(self, checked):
        if not self.contours:
            self.ui.actionMeasureScale.setChecked(False)
            return
        print('Measure Real World Scale')
        if checked:
            cv2.namedWindow('live show', cv2.WINDOW_NORMAL)
            cv2.setMouseCallback('result', self.mouse_callback_measureScale)
        else:
            global tmp_small_img
            tmp_small_img = None
            cv2.destroyWindow('live show')
            cv2.setMouseCallback('result', lambda *args: None)

    def ImportScaleFromImage(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "Select Image")
        excel_filepath = f'{os.path.join(os.path.dirname(filepath), os.path.split(os.path.dirname(filepath))[1])}.xlsx'
        filename = os.path.split(filepath)[1]
        if not os.path.exists(excel_filepath):
            QMessageBox.warning(self, 'Error', 'Scale Not Found!')
            return

        workbook = openpyxl.load_workbook(excel_filepath)
        worksheet = workbook.worksheets[0]
        for row in worksheet.iter_rows():
            if row[0].value == filename:
                self.actual_scale = row[1].value
                self.ui.label_actual_scale.setText(f'ActualScale (um/pix): {self.actual_scale}')
                QMessageBox.information(self, filename, f'Set Scale to {self.actual_scale}')
                return
        QMessageBox.warning(self, 'Error', 'Scale Not Found!')

    # ------------------------mouse callbacks--------------------------------
    def mouse_callback_modify(self, event, x, y, flags, param):
        # 修正模式的鼠标回调函数
        # 左键点击-拖动可以分割轮廓
        # 右键点击可以删除轮廓, 优先删除最大面积的轮廓
        x = round(x / self.show_scale)
        y = round(y / self.show_scale)
        global flag_draw, ix, iy, tmp_contour, tmp_img, current_contour, nearest_contour_idx
        if event == cv2.EVENT_LBUTTONDOWN:
            nearest_contour_idx, min_dist = utils.findNearestContour(self.contour_history[-1], x, y)
            current_contour = self.contour_history[-1][nearest_contour_idx]
            if (abs(min_dist) < self.lineWidth):
                flag_draw = True
                ix, iy = x, y
                tmp_contour = []
                nearest_pt_idx, nearest_pt = utils.findNearestPoint_on_contour(x, y, current_contour, )
                tmp_contour.append(nearest_pt)
                tmp_img = utils.draw_contours(self.currentImg, self.contours_filter(self.contour_history[-1]),
                                              line_width=self.lineWidth, selected_idx=-1)
                cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))
            else:
                return
        elif event == cv2.EVENT_MOUSEMOVE:
            _, min_dist = utils.findNearestContour(self.contour_history[-1], x, y)
            if (abs(min_dist) < self.lineWidth):
                win32api.SetCursor(win32api.LoadCursor(0, win32con.IDC_HAND))
            else:
                win32api.SetCursor(win32api.LoadCursor(0, win32con.IDC_CROSS))
            if flag_draw:
                tmp_contour.append([x, y])
                cv2.line(tmp_img, (ix, iy), (x, y), (0, 0, 255), self.lineWidth)
                ix, iy = x, y
                cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))
        elif event == cv2.EVENT_LBUTTONUP:
            flag_draw = False
            if not tmp_contour:
                return
            tmp_contour.append([x, y])
            nearest_pt_idx, nearest_pt = utils.findNearestPoint_on_contour(x, y, current_contour, )
            tmp_contour.append(nearest_pt)
            # cv2.line(tmp_img, (ix, iy), (x, y), (0, 0, 255), 2)
            ix, iy = x, y
            self.contour_history.append(utils.split_contour(self.contour_history[-1], nearest_contour_idx, tmp_contour))
            self.show_result()
            # cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))
        elif event == cv2.EVENT_RBUTTONUP:
            # 删除点
            # 规则如下：
            # 在轮廓上，即鼠标为手型时，删除鼠标最近的轮廓
            # 在轮廓内时，删除包围点的面积最大的轮廓
            # idx, min_dist = utils.findNearestContour(self.contour_history[-1], x, y)
            # if min_dist > self.lineWidth:
            #     return

            idx, min_dist = utils.findNearestContour(self.contour_history[-1], x, y)
            if abs(min_dist) > self.lineWidth:
                # return
                bounding_contours_idx = utils.findBoundingContours(self.contour_history[-1], x, y)
                if not bounding_contours_idx:
                    return
                idx = np.argmax([cv2.contourArea(self.contour_history[-1][i]) for i in bounding_contours_idx])
                idx = bounding_contours_idx[idx]

            contours = self.contour_history[-1].copy()
            del contours[idx]
            self.contour_history.append(contours)
            tmp_img = utils.draw_contours(self.currentImg, self.contours_filter(self.contour_history[-1]),
                                          line_width=self.lineWidth, selected_idx=-1)
            cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))

        elif event == cv2.EVENT_MBUTTONDOWN:
            # 撤销
            if len(self.contour_history) == 1:
                return
            del self.contour_history[-1]
            self.show_result()

    def mouse_callback_measureScale(self, event, x, y, flags, param):
        global flag_measuring, tmp_measure_start_x, tmp_measure_end_x, tmp_y, tmp_img, tmp_padding_img
        x = round(x / self.show_scale)
        y = round(y / self.show_scale)
        length = 100
        lx = x - length
        rx = x + length
        ly = y - length
        ry = y + length
        if tmp_padding_img is None:
            tmp_padding_img = cv2.copyMakeBorder(self.currentImg, length, length, length, length, cv2.BORDER_CONSTANT,
                                                 value=(0, 0, 0))
        if flag_measuring:
            tmp_measure_end_x = x
        if event == cv2.EVENT_LBUTTONDOWN:
            flag_measuring = True
            tmp_measure_start_x = x
            tmp_y = y
        elif event == cv2.EVENT_LBUTTONUP:
            flag_measuring = False
            print(f'{tmp_measure_start_x}-{tmp_measure_end_x}')
            dialog = ActualScaleDialog(self)
            num = dialog.get_data()
            if num is not None:
                result = QMessageBox.question(self, 'Confirmation', f"You entered {num}, Are you sure?",
                                              QMessageBox.Ok | QMessageBox.Cancel)
                if result == QMessageBox.Ok:
                    print(f'{tmp_measure_start_x}-{tmp_measure_end_x}/{num}')
                    self.actual_scale = num / abs(tmp_measure_end_x - tmp_measure_start_x)
                    self.ui.label_actual_scale.setText(f'ActualScale (um/pix): {self.actual_scale}')
                    self.ui.actionMeasureScale.setChecked(False)
                    print(f"You entered {num}, set actual_scale {self.actual_scale}")

            self.show_result()

        if flag_measuring:
            tmp_img = cv2.line(self.currentImg.copy(), (tmp_measure_start_x, tmp_y), (x, tmp_y),
                               (0, 0, 255), self.lineWidth)
            cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))

        tmp_small_img = tmp_padding_img[ly + length:ry + length + 1, lx + length:rx + length + 1].copy()
        cv2.circle(tmp_small_img, (length, length), 5, (0, 255, 0), -1)
        cv2.imshow('live show', tmp_small_img)

    # -----------------------------functions-----------------------------------
    def read_img(self, ):
        if self.currentImgIdx == self.formerImgIdx:
            return
        self.formerImgIdx = self.currentImgIdx
        img_path = os.path.join(self.currentFolder, self.imgList[self.currentImgIdx])
        self.currentImg = cv2.imread(img_path)

    def contours_filter(self, contours):

        # 面积大小筛选
        contours = [c for c in contours if self.currentImgAreas[self.minAreaIndex] <= cv2.contourArea(c)]
        # 删除边缘
        if self.ui.checkBox_delete_edge_contour.isChecked():
            contours = utils.DeleteEdgeContour(contours, self.currentImg.shape[1], self.currentImg.shape[0])
        return contours

    def show_result(self, ):
        if self.FLAG_modify_mode:
            scale = self.show_scale
            contours_filtered = self.contours_filter(self.contour_history[-1])  # 展示history最后一个
            contours = utils.scale_contour_ToShow(contours_filtered, scale=scale)
            img = utils.resize_img(self.currentImg, scale=self.show_scale)
            img_drawed_scaled = utils.draw_contours(img.copy(), contours, line_width=int(10 * scale))
        else:
            scale = self.show_scale
            contours_filtered = self.contours_filter(self.contours)
            contours = utils.scale_contour_ToShow(contours_filtered, scale=scale)
            img = utils.resize_img(self.currentImg, scale=self.show_scale)
            img_drawed_scaled = utils.draw_contours(img.copy(), contours, line_width=int(10 * scale))

        if img_drawed_scaled is None:
            return
        if self.ui.checkBox_draw_contours.isChecked():
            cv2.imshow('result', img_drawed_scaled)
        else:
            cv2.imshow('result', img)


# 调整轮廓 临时变量
flag_draw = False
tmp_contour = []
# 测量真实比例尺 临时变量
flag_measuring = False
tmp_padding_img = None
if __name__ == '__main__':
    app = QApplication(sys.argv)
    MainWindow = MainWindow()
    MainWindow.show()
    sys.exit(app.exec_())
