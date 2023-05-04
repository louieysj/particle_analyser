import os
import sys
import time
import cv2
import random
import numpy as np
import win32api, win32con
import openpyxl
from PIL import Image
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QDialog, QMainWindow, QMessageBox, QFileDialog, QLabel
from PyQt5.uic import loadUi

import detector
from UI.main_window_ui import Ui_MainWindow
from UI.measure_scale_dialog import ActualScaleDialog
import utils


class MainWindow(QMainWindow):
    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        # 1. 直接load ui:
        # loadUi(r'UI\MainWindow.ui', self)
        # 2. 先 pyuic5  UI/MainWindow.ui -o UI/main_window_ui.py, 后load
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.status_label = QLabel('')
        self.ui.statusbar.addWidget(self.status_label)
        self.ui.slider_depth.setPageStep(5)
        # self.ui.slider_depth.setTracking(False)

        # ----------参数初始化-------------
        self.currentFolder = None  # 当前文件夹
        self.imgList = None  # 当前文件夹下的图片文件名
        self.formerImgIdx = None
        self.currentImgIdx = 0  # imgList的当前index
        self.currentImg = None
        self.currentImgDepth = None
        self.contours = None  # 当天打开的img的检测到的所有轮廓，原始数据
        self.show_scale = self.ui.slider_scale.value() / 100  # cv2.imshow的缩放比例
        self.actual_scale = None  # 真实世界的比例尺 um/每像素
        self.minAreaIndex = None  # 筛选面积大小
        self.currentImgAreas = None
        self.DepthThresh = None  # 筛选深度大小
        self.lineWidth = 15

        self.detector = detector.Detector()

        self.FLAG_modify_mode = False

        # 绑定按钮
        self.connectSlots()
        # 修改自定义组件

        # debug
        # self.ui.RangeSlider_area.setData([random.randint(0, 50) for _ in range(50)])
        # self.currentFolder = './3d data'
        # self.imgList = os.listdir(self.currentFolder)
        # self.currentImgIdx = 3
        # self.formerImgIdx = None
        # self.ui.label_current_folder.setText(self.currentFolder)
        # self.ui.label_current_file.setText(
        #     f'({self.currentImgIdx + 1}/{len(self.imgList)}) [{self.imgList[self.currentImgIdx]}]')
        # img_path = os.path.join(self.currentFolder, self.imgList[self.currentImgIdx])

        # self.currentImg = cv2.imread(img_path)
        # self.currentImgDepth = utils.read_depth_from_csv(os.path.splitext(img_path)[0] + '.csv')
        # self.detect_contour()

    def connectSlots(self):
        # buttons
        self.ui.btn_goto_previous.clicked.connect(self.goto_previous)
        self.ui.btn_goto_next.clicked.connect(self.goto_next)
        self.ui.btn_detect_contour.clicked.connect(self.detect_contour)

        # others
        self.ui.slider_scale.valueChanged.connect(self.scale_OnChanged)
        self.ui.slider_depth.valueChanged.connect(self.depth_range_changed)
        self.ui.slider_area.valueChanged.connect(self.area_range_changed)
        self.ui.checkBox_delete_edge_contour.stateChanged.connect(self.show_result)
        self.ui.checkBox_draw_contours.stateChanged.connect(self.show_result)

        # actions
        self.ui.actionOpenFolder.triggered.connect(self.OpenFolder)
        self.ui.actionModifyMode.triggered.connect(self.switchModifyMode)
        self.ui.actionMeasureScale.triggered.connect(self.measureActualScaleMod)
        self.ui.actionPreprocessDepth.triggered.connect(self.preprocessDepthData)
        self.ui.actionExportData.triggered.connect(self.ExportData)

        # 自定义组件
        # self.ui.RangeSlider_area.rangeChanged.connect(self.area_range_changed)

    ################################ Slots ################################
    # --------------------------pushbutoon slots---------------------------
    def detect_contour(self):
        if not self.imgList:
            QMessageBox.information(self, 'info', 'Open Folder First!')
            return

        img_path = os.path.join(self.currentFolder, self.imgList[self.currentImgIdx])
        csv_path = os.path.splitext(img_path)[0] + '.csv'
        depth_img_path = os.path.join(self.currentFolder, 'PREPROCESSED_DEPTH',
                                      os.path.splitext(self.imgList[self.currentImgIdx])[0] + '.png')
        if os.path.exists(depth_img_path):
            pass
        elif os.path.exists(csv_path):
            result = QMessageBox.warning(self, 'Confirmation',
                                         f"Detect without preprocess depth data will cause slow response,\n"
                                         f"Continue anyway?",
                                         QMessageBox.Ok | QMessageBox.Cancel)
            if result == QMessageBox.Cancel:
                return
        else:
            result = QMessageBox.warning(self, 'Confirmation',
                                         f"CSV file not found, Continue detect without depth data?",
                                         QMessageBox.Ok | QMessageBox.Cancel)
            if result == QMessageBox.Cancel:
                return

        self.read_img_and_depth()

        print('detect_contour')
        st = time.time()
        self.contours = self.detector.detect(self.currentImg, self.currentImgDepth, self.DepthThresh)
        self.contours = self.split_boarder_contour(self.contours)
        print(time.time() - st)

        areas = [cv2.contourArea(c) for c in self.contours]
        self.currentImgAreas = sorted(areas)
        self.minAreaIndex = int(len(areas) * 0.2)
        self.ui.slider_area.setRange(0, len(areas) - 1)
        self.ui.slider_area.setValue(self.minAreaIndex)

        self.show_result()
        cv2.waitKey()

    def goto_next(self):
        if not self.imgList:
            return
        self.currentImgIdx = min(len(self.imgList) - 1, self.currentImgIdx + 1)
        self.ui.label_current_file.setText(
            f'({self.currentImgIdx + 1}/{len(self.imgList)}) [{self.imgList[self.currentImgIdx]}]')
        # self.read_img_and_depth()

        # self.detect_contour()

    def goto_previous(self):
        if not self.imgList:
            return
        self.currentImgIdx = max(0, self.currentImgIdx - 1)
        self.ui.label_current_file.setText(
            f'({self.currentImgIdx + 1}/{len(self.imgList)}) [{self.imgList[self.currentImgIdx]}]')
        # self.read_img_and_depth()

        # self.detect_contour()

    # ---------------------------slider slots-------------------------------
    def scale_OnChanged(self, ):
        self.show_scale = self.ui.slider_scale.value() / 100
        self.ui.label_scale.setText(f'ShowScale:{self.show_scale:.2f}')
        if self.contours:
            self.show_result()

    def depth_range_changed(self, ):
        # print(min_area, max_area)
        if self.currentImgDepth is None:
            return
        min_depth, max_depth = self.ui.slider_depth.minimum(), self.ui.slider_depth.maximum()
        self.DepthThresh = self.ui.slider_depth.value()
        self.ui.label_minmaxDepth.setText(
            f'DepthFilter(minDepth/threshold/maxDepth)：{min_depth}/{self.DepthThresh}/{max_depth}')
        if self.contours:
            self.detect_contour()

    # ---------------------------action slots-------------------------------
    def OpenFolder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select Directory")  # 打开文件夹选择对话框
        if folder_path:
            imgList = os.listdir(folder_path)
            allow_ext = ['jpg', 'png', 'tif', 'bmp']
            imgList = [i for i in imgList if
                       i.lower().split('.')[-1] in allow_ext and not i.startswith('PREPROCESSED_')]
            if imgList:
                self.currentFolder = folder_path
                self.imgList = imgList
                self.currentImgIdx = 0
                self.formerImgIdx = None
                # self.read_img_and_depth()

                self.ui.label_current_folder.setText(self.currentFolder)
                self.ui.label_current_file.setText(
                    f'({self.currentImgIdx + 1}/{len(self.imgList)}) [{self.imgList[self.currentImgIdx]}]')
            else:
                QMessageBox.information(self, 'info', 'There is no image in the folder!')

    def ExportData(self):
        if not self.actual_scale:
            QMessageBox.information(self, 'info', 'Please set scale first!')
            return
        if self.FLAG_modify_mode:
            QMessageBox.information(self, 'info', "You haven't quit modify mode!")
            return

        print('ExportData')
        try:
            excel_filename = f'{os.path.basename(self.currentFolder)}.xlsx'
            excel_filename = os.path.join(self.currentFolder, excel_filename)
            if os.path.exists(excel_filename):
                workbook = openpyxl.load_workbook(excel_filename)
            else:
                workbook = openpyxl.Workbook()
            if self.imgList[self.currentImgIdx] in workbook.sheetnames:
                del workbook[self.imgList[self.currentImgIdx]]
                # worksheet.delete_rows(1, worksheet.max_row)
                # worksheet.delete_cols(1, worksheet.max_column)

            # 写入数据
            worksheet = workbook.create_sheet(self.imgList[self.currentImgIdx])
            worksheet.append(['Circularity', 'Area', 'Diameter'])
            diameters = []
            for contour in self.contours_filter(self.contours):
                rescaled = (contour * self.actual_scale).astype(np.float32)
                Circularity = utils.cal_circularity(rescaled)
                Diameter = utils.cal_diameter(rescaled)
                diameters.append(Diameter)
                Area = utils.cal_area(rescaled)
                worksheet.append([Circularity, Area, Diameter, ])

            # 写入比例尺
            worksheet = workbook.worksheets[0]
            worksheet.append([self.imgList[self.currentImgIdx], self.actual_scale])

            # 显示颗粒总数量和d97, d50
            total_num = len(diameters)
            sorted_diameters = sorted(diameters)
            d97 = sorted_diameters[int((total_num - 1) * 0.97)]
            d50 = sorted_diameters[int((total_num - 1) * 0.5)]
            self.ui.statusbar.showMessage(f"TotalNum:{total_num}, d97:{d97:.3f}, d50:{d50:.3f}")

            workbook.save(excel_filename)
            print('Excel saved!')
            QMessageBox.information(self, 'info', 'Saved!')
        except Exception as e:
            print(e)
            QMessageBox.warning(self, 'Error Happened!!!', str(e))

    def switchModifyMode(self, checked):
        if not self.contours:
            self.ui.actionModifyMode.setChecked(False)
            return
        print('switch to edit Mode')
        if checked:
            QMessageBox.information(self,
                                    'ModifyMode',
                                    'Press the left button to start drawing, release it to end drawing;\n'
                                    'Right click to delete: If it is a hand shape, delete the contour at the current position, '
                                    'otherwise delete the largest contour surrounding the current position;\n'
                                    'Middle-click to undo'
                                    )
        if checked:
            self.FLAG_modify_mode = True
            # cv2.namedWindow('modify', cv2.WINDOW_NORMAL)
            cv2.setMouseCallback('result', self.mouse_callback_modify)
            self.contour_history = [self.contours]  # 临时
        else:
            self.FLAG_modify_mode = False
            # cv2.destroyWindow('modify')
            cv2.setMouseCallback('result', lambda *args: None)
            self.contours = self.contour_history[-1]
        self.show_result()

    def measureActualScaleMod(self, checked):
        if not self.contours:
            self.ui.actionMeasureScale.setChecked(False)
            return
        print('Measure Real World Scale')
        global tmp_padding_img
        tmp_padding_img = None
        if checked:
            cv2.namedWindow('live show', cv2.WINDOW_NORMAL)
            cv2.setMouseCallback('result', self.mouse_callback_measureScale)
        else:
            cv2.destroyWindow('live show')
            cv2.setMouseCallback('result', lambda *args: None)

    def preprocessDepthData(self, checked):
        if not self.currentFolder:
            return
        for img_filename in self.imgList:
            self.status_label.setText(f"正在预处理 {img_filename}")
            img_path = os.path.join(self.currentFolder, img_filename)
            csv_path = os.path.splitext(img_path)[0] + '.csv'
            depth_img_path = os.path.join(self.currentFolder, 'PREPROCESSED_DEPTH',
                                          os.path.splitext(img_filename)[0] + '.png')
            if not os.path.exists(os.path.join(self.currentFolder, 'PREPROCESSED_DEPTH')):
                os.mkdir(os.path.join(self.currentFolder, 'PREPROCESSED_DEPTH'))
            if os.path.exists(depth_img_path):
                continue
            depth_data = np.loadtxt(csv_path, delimiter=',')
            depth_data = utils.replace_minus1(depth_data)
            depth_img = Image.fromarray(depth_data.astype(np.int16))
            depth_img.save(depth_img_path)
        self.status_label.setText('预处理完成')

    # ----------------------custom signal slots-----------------------------
    def area_range_changed(self, min_area_idx):
        # print(min_area, max_area)
        if not self.contours:
            return
        self.minAreaIndex = min_area_idx
        min_area = min(self.currentImgAreas)
        max_area = max(self.currentImgAreas)
        self.ui.label_minmaxArea.setText(
            f'MinArea:{min_area:.1f}, MaxArea:{max_area:.1f}\n'
            f'Filter MinArea:{self.currentImgAreas[min_area_idx]:.1f}'
        )
        self.show_result()

    # ------------------------mouse callbacks--------------------------------
    def mouse_callback_modify(self, event, x, y, flags, param):
        # 修正模式的鼠标回调函数
        # 左键点击-拖动可以分割轮廓
        # 右键点击可以删除轮廓, 优先删除最大面积的轮廓
        x = round(x / self.show_scale)
        y = round(y / self.show_scale)
        global flag_draw, ix, iy, tmp_contour, tmp_img, current_contour, nearest_contour_idx
        if event == cv2.EVENT_LBUTTONDOWN:
            nearest_contour_idx, min_dist = utils.findNearestContour(self.contour_history[-1], x, y)
            current_contour = self.contour_history[-1][nearest_contour_idx]
            if (abs(min_dist) < self.lineWidth):
                flag_draw = True
                ix, iy = x, y
                tmp_contour = []
                nearest_pt_idx, nearest_pt = utils.findNearestPoint_on_contour(x, y, current_contour, )
                tmp_contour.append(nearest_pt)
                tmp_img = utils.draw_contours(self.currentImg, self.contours_filter(self.contour_history[-1]),
                                              line_width=self.lineWidth, selected_idx=-1)
                cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))
            else:
                return
        elif event == cv2.EVENT_MOUSEMOVE:
            _, min_dist = utils.findNearestContour(self.contour_history[-1], x, y)
            if (abs(min_dist) < self.lineWidth):
                win32api.SetCursor(win32api.LoadCursor(0, win32con.IDC_HAND))
            else:
                win32api.SetCursor(win32api.LoadCursor(0, win32con.IDC_CROSS))
            if flag_draw:
                tmp_contour.append([x, y])
                cv2.line(tmp_img, (ix, iy), (x, y), (0, 0, 255), self.lineWidth)
                ix, iy = x, y
                cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))
        elif event == cv2.EVENT_LBUTTONUP:
            try:
                flag_draw = False
                if not tmp_contour:
                    return
                tmp_contour.append([x, y])
                nearest_pt_idx, nearest_pt = utils.findNearestPoint_on_contour(x, y, current_contour, )
                tmp_contour.append(nearest_pt)
                # cv2.line(tmp_img, (ix, iy), (x, y), (0, 0, 255), 2)
                ix, iy = x, y
                self.contour_history.append(utils.split_contour(self.contour_history[-1], nearest_contour_idx, tmp_contour))
            except:
                pass
            self.show_result()
            # cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))
        elif event == cv2.EVENT_RBUTTONUP:
            # 删除点
            # 规则如下：
            # 在轮廓上，即鼠标为手型时，删除鼠标最近的轮廓
            # 在轮廓内时，删除包围点的面积最大的轮廓
            # idx, min_dist = utils.findNearestContour(self.contour_history[-1], x, y)
            # if min_dist > self.lineWidth:
            #     return

            idx, min_dist = utils.findNearestContour(self.contour_history[-1], x, y)
            if abs(min_dist) > self.lineWidth:
                # return
                bounding_contours_idx = utils.findBoundingContours(self.contour_history[-1], x, y)
                if not bounding_contours_idx:
                    return
                idx = np.argmax([cv2.contourArea(self.contour_history[-1][i]) for i in bounding_contours_idx])
                idx = bounding_contours_idx[idx]

            contours = self.contour_history[-1].copy()
            del contours[idx]
            self.contour_history.append(contours)
            tmp_img = utils.draw_contours(self.currentImg, self.contours_filter(self.contour_history[-1]),
                                          line_width=self.lineWidth, selected_idx=-1)
            cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))

        elif event == cv2.EVENT_MBUTTONDOWN:
            # 撤销
            if len(self.contour_history) == 1:
                return
            del self.contour_history[-1]
            self.show_result()

    def mouse_callback_measureScale(self, event, x, y, flags, param):
        global flag_measuring, tmp_measure_start_x, tmp_measure_end_x, tmp_y, tmp_img, tmp_padding_img
        x = round(x / self.show_scale)
        y = round(y / self.show_scale)
        length = 100
        lx = x - length
        rx = x + length
        ly = y - length
        ry = y + length
        if tmp_padding_img is None:
            tmp_padding_img = cv2.copyMakeBorder(self.currentImg, length, length, length, length, cv2.BORDER_CONSTANT,
                                                 value=(0, 0, 0))
        if flag_measuring:
            tmp_measure_end_x = x
        if event == cv2.EVENT_LBUTTONDOWN:
            flag_measuring = True
            tmp_measure_start_x = x
            tmp_y = y
        elif event == cv2.EVENT_LBUTTONUP:
            flag_measuring = False
            print(f'{tmp_measure_start_x}-{tmp_measure_end_x}')
            dialog = ActualScaleDialog(self)
            num = dialog.get_data()
            if num is not None:
                result = QMessageBox.question(self, 'Confirmation', f"You entered {num}, Are you sure?",
                                              QMessageBox.Ok | QMessageBox.Cancel)
                if result == QMessageBox.Ok:
                    print(f'{tmp_measure_start_x}-{tmp_measure_end_x}/{num}')
                    self.actual_scale = num / abs(tmp_measure_end_x - tmp_measure_start_x)
                    self.ui.label_actual_scale.setText(f'ActualScale (um/pix): {self.actual_scale}')
                    self.ui.actionMeasureScale.setChecked(False)
                    print(f"You entered {num}, set actual_scale {self.actual_scale}")

            self.show_result()

        if flag_measuring:
            tmp_img = cv2.line(self.currentImg.copy(), (tmp_measure_start_x, tmp_y), (x, tmp_y),
                               (0, 0, 255), self.lineWidth)
            cv2.imshow('result', utils.resize_img(tmp_img, scale=self.show_scale))

        tmp_small_img = tmp_padding_img[ly + length:ry + length + 1, lx + length:rx + length + 1].copy()
        cv2.circle(tmp_small_img, (length, length), 5, (0, 255, 0), -1)
        cv2.imshow('live show', tmp_small_img)

    # -----------------------------functions-----------------------------------
    def read_img_and_depth(self, ):
        img_path = os.path.join(self.currentFolder, self.imgList[self.currentImgIdx])
        csv_path = os.path.splitext(img_path)[0] + '.csv'
        depth_img_path = os.path.join(self.currentFolder, 'PREPROCESSED_DEPTH',
                                      os.path.splitext(self.imgList[self.currentImgIdx])[0] + '.png')

        # if os.path.exists(depth_img_path):
        #     pass
        # elif os.path.exists(csv_path):
        #     result = QMessageBox.warning(self, 'Confirmation',
        #         f"Detect without preprocess depth data will cause slow response,\n"
        #         f"Continue anyway?",
        #         QMessageBox.Ok | QMessageBox.Cancel)
        #     if result == QMessageBox.Cancel:
        #         return False
        # else:
        #     result = QMessageBox.warning(self, 'Confirmation', f"CSV file not found, Continue detect without depth data?",
        #                                   QMessageBox.Ok | QMessageBox.Cancel)
        #     if result == QMessageBox.Cancel:
        #         return False

        if self.currentImgIdx == self.formerImgIdx:
            return
        self.formerImgIdx = self.currentImgIdx

        self.currentImg = cv2.imread(img_path)
        if os.path.exists(depth_img_path):
            self.currentImgDepth = cv2.imread(depth_img_path, cv2.IMREAD_ANYDEPTH).astype('int16')
        elif os.path.exists(csv_path):
            self.currentImgDepth = utils.read_depth_from_csv(csv_path)
        else:
            self.currentImgDepth = None
            self.ui.label_minmaxDepth.setText("Depth data not found")
            return
        min_depth, max_depth = self.currentImgDepth.min(), self.currentImgDepth.max()
        self.DepthThresh = min_depth + (max_depth - min_depth) / 2
        self.ui.slider_depth.setRange(int(min_depth), int(max_depth))
        self.ui.slider_depth.setValue(int(self.DepthThresh))
        self.ui.label_minmaxDepth.setText(
            f'深度筛选(最小深度/阈值深度/最大深度)：{min_depth}/{self.DepthThresh}/{max_depth}')
        return

    def contours_filter(self, contours):
        # 1.面积大小筛选
        contours = [c for c in contours if self.currentImgAreas[self.minAreaIndex] <= cv2.contourArea(c)]
        # 2.删除边缘
        if self.ui.checkBox_delete_edge_contour.isChecked():
            contours = utils.DeleteEdgeContour(contours, self.currentImg.shape[1], self.currentImg.shape[0])
        return contours

    def split_boarder_contour(self, contours):
        # 处理掉边缘的轮廓，因为边缘的轮廓是反向包围整个画面的，所以要遍历边缘，把边缘的点取反，最后拆分
        contours = list(contours)
        areas = [cv2.contourArea(i) for i in contours]
        if np.max(areas) <= 0.5 * self.currentImg.shape[0] * self.currentImg.shape[1]:
            return contours

        def check_crossing_vertex(h, w, p1, p2):
            x1, y1 = p1
            x2, y2 = p2
            line1 = [
                (x1 if x1 in (0, w - 1) else 0, y1 if y1 in (0, h - 1) else 0),
                (x1 if x1 in (0, w - 1) else w - 1, y1 if y1 in (0, h - 1) else h - 1)
            ]
            line2 = [
                (x2 if x2 in (0, w - 1) else 0, y2 if y2 in (0, h - 1) else 0),
                (x2 if x2 in (0, w - 1) else w - 1, y2 if y2 in (0, h - 1) else h - 1)
            ]
            if line1 == line2:
                return None
            else:
                for p11 in line1:
                    for p22 in line2:
                        if p11 == p22:
                            return p11

        max_contour = contours.pop(np.argmax(areas))
        h, w, c = self.currentImg.shape
        # 遍历contour，分离
        new_contours = []
        flag = False
        i = 0;
        j = 0
        for idx, point in enumerate(max_contour):
            x, y = point[0]
            j = idx
            if x not in (0, w - 1) and y not in (0, h - 1):  # 点x,y不在四周的边界上
                if not flag:
                    flag = True
                continue
            else:  # 点x,y在四周的边界上
                if not flag:
                    i = idx
                    continue
                else:
                    flag = False
                    # 分离轮廓[i:j]
                    tmp = max_contour[i:j + 1]
                    # 处理跨越四个顶点的情况，加上顶点的坐标
                    x1, y1 = max_contour[i][0]
                    x2, y2 = max_contour[j][0]
                    vertex = check_crossing_vertex(h, w, (x1, y1), (x2, y2))
                    if vertex:
                        tmp = np.append(tmp, np.array(vertex).reshape((1, 1, 2)), axis=0)
                    new_contours.append(tmp)
                    i = idx

        contours += new_contours
        return contours

    def show_result(self, ):
        if self.FLAG_modify_mode:
            scale = self.show_scale
            contours_filtered = self.contours_filter(self.contour_history[-1])  # 展示history最后一个
            contours = utils.scale_contour_ToShow(contours_filtered, scale=scale)
            img = utils.resize_img(self.currentImg, scale=self.show_scale)
            img_drawed_scaled = utils.draw_contours(img.copy(), contours, line_width=int(10 * scale))
        else:
            scale = self.show_scale
            contours_filtered = self.contours_filter(self.contours)
            contours = utils.scale_contour_ToShow(contours_filtered, scale=scale)
            img = utils.resize_img(self.currentImg, scale=self.show_scale)
            img_drawed_scaled = utils.draw_contours(img.copy(), contours, line_width=int(10 * scale))

        if img_drawed_scaled is None:
            return
        if self.ui.checkBox_draw_contours.isChecked():
            cv2.imshow('result', img_drawed_scaled)
        else:
            cv2.imshow('result', img)


# 调整轮廓 临时变量
flag_draw = False
tmp_contour = []
# 测量真实比例尺 临时变量
flag_measuring = False
tmp_padding_img = None
if __name__ == '__main__':
    app = QApplication(sys.argv)
    MainWindow = MainWindow()
    MainWindow.show()
    sys.exit(app.exec_())
